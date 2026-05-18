"""Read refrigeration-equipment energy specs from CSV / XLSX and produce
a per-(koneikko, laitetunnus) lookup that the orchestrator can merge
into FI_Tekninen.

The Finnish refrigeration workflow keeps energy specs (Jäähdytysteho,
Sähköteho, Kylmäaine, ilmavirta, äänitiedot, käyttölämpötilat) in an
Excel spreadsheet — one row per höyrystin / lauhdutin / kompressori,
keyed by koneikkotunnus + laitetunnus. POSITIO-blokit DXF-pohjassa
kantavat samat tunnukset, joten kun POSITIO-linkitys on kerran tehty,
me voidaan automaattisesti hakea kunkin laitteen oikeat tehotiedot
listalta ja kirjoittaa ne IFC:n FI_Tekninen-PSetiin.

Sarakeotsikoiden mätsäys on tarkoituksella laaja:
``Jäähdytysteho [kW]`` / ``Q_kW`` / ``Cooling capacity`` kaikki
mappautuvat samaksi FI_Tekninen-kentäksi ``Jäähdytysteho``.
Yhdistetyt otsikot kuten ``KYLMÄ-/SÄHKÖ-/VASTUSTEHO [kW]``
laajennetaan kolmeksi sarakkeeksi (Jäähdytysteho/Sähköteho/Vastusteho).

Lauri:n RefDesign-pohjat ryhmittelevät rivejä koneikon mukaan
sektio-otsikoilla (esim. ``PAKASTEET JK1``) ja jättävät
``REV.``-sarakkeen tyhjäksi useimmilla data-riveillä. Lukija
forward-fillaa koneikon: viimeinen ei-tyhjä REV.-arvo TAI sektio-
otsikon tunnistama JKx/RKx/KKx -koodi periytyy seuraaville riveille.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable


# Canonical FI_Tekninen field name → list of recognised aliases.
# Matching is case-insensitive against the column header after stripping
# whitespace, brackets, and the unit ("[kW]", "(kW)" etc.).
_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "Jäähdytysteho": (
        "jäähdytysteho",
        "jaahdytysteho",
        "jäähdytys",
        "jaahdytys",
        "jäähdytys teho",
        "kylmäteho",
        "cooling capacity",
        "cooling power",
        "q_kw",
        "q kw",
        "qe",
        "qe_kw",
    ),
    "Lauhdutusteho": (
        "lauhdutusteho",
        "lauhdutus teho",
        "lauhdutus",
        "qc",
        "qc_kw",
        "condensing capacity",
        "condenser power",
    ),
    "Sähköteho": (
        "sähköteho",
        "sahkoteho",
        "sähkö teho",
        "sähkö",
        "sahko",
        "p_kw",
        "p kw",
        "electric power",
        "electrical power",
    ),
    "Kylmäaine": (
        "kylmäaine",
        "kylmaaine",
        "kylmä aine",
        "kylma aine",
        "refrigerant",
        "aine",
    ),
    "Ilmavirta": (
        "ilmavirta",
        "ilma virta",
        "air flow",
        "airflow",
        "air volume",
        "m3/h",
        "m³/h",
    ),
    "Ääniteho": (
        "ääniteho",
        "aaniteho",
        "ääni teho",
        "äänitaso",
        "aanitaso",
        "sound power",
        "noise level",
        "db(a)",
        "lwa",
    ),
    "Käyttölämpötila": (
        "käyttölämpötila",
        "kayttolampotila",
        "käyttö lämpötila",
        "operating temperature",
        "lämpötila",
        "lampotila",
    ),
    "Höyrystymislämpötila": (
        "höyrystymislämpötila",
        "hoyrystymislampotila",
        "evaporation temperature",
        "te",
        "to",
    ),
    "Lauhtumislämpötila": (
        "lauhtumislämpötila",
        "lauhtumislampotila",
        "condensing temperature",
        "tc",
    ),
    "Vastusteho": (
        "vastusteho",
        "sulatusteho",
        "vastus teho",
        "defrost",
        "defrost power",
        "heater",
    ),
    "Jännite": (
        "jännite",
        "jannite",
        "voltage",
        "u_v",
    ),
    "Jäähdyttävä vaikutus": (
        "jäähdyttävä vaikutus",
        "jaahdyttava vaikutus",
        "cooling effect",
    ),
}

# Header aliases for the two key columns. Same case-insensitive matching.
_KONEIKKO_ALIASES: tuple[str, ...] = (
    "koneikko",
    "koneikkotunnus",
    "koneikko-tunnus",
    "koneikko tunnus",
    "koneryhmä",
    "koneryhma",
    "unit",
    "system",
    # RefDesign-konventio Lauri'n teholuetteloissa: "REV." -sarake
    # sisältää koneikkotunnuksen (JK1, JK2…), ei revision-numeroa.
    "rev.",
    "rev",
)
_LAITETUNNUS_ALIASES: tuple[str, ...] = (
    "laitetunnus",
    "laite tunnus",
    "laite-tunnus",
    "laite",
    "tunnus",
    "pos",
    "pos.",
    "positio",
    "positionumero",
    "numero",
    "nr",
    "device",
    "tag",
)


@dataclass(frozen=True)
class EnergySpec:
    """Resolved energy fields for one device. ``fields`` is a
    canonical-name → string-value mapping ready to merge into
    FI_Tekninen."""

    koneikko: str
    laitetunnus: str
    fields: dict[str, str]


def _normalise_header(raw: str) -> str:
    """Lowercase + strip + remove unit brackets / annotations.

    "Jäähdytysteho [kW]" → "jäähdytysteho"
    "Q_kW (cooling)"     → "q_kw"
    "  Laitetunnus  "    → "laitetunnus"
    """
    s = raw or ""
    # Drop everything from the first '(' or '[' onward — that's almost
    # always a unit annotation we already strip implicitly.
    for sep in ("[", "("):
        idx = s.find(sep)
        if idx != -1:
            s = s[:idx]
    return s.strip().lower()


def _match_alias(header: str, aliases: Iterable[str]) -> bool:
    """Return True when the normalised header equals OR contains any alias.

    Short aliases (≤3 characters) require an exact match — substring
    matching for tokens like "te", "tc", "u_v" would otherwise leak
    across distinct fields ("vastusteho" should not match "Te" =
    Evaporation temperature).
    """
    h = _normalise_header(header)
    for alias in aliases:
        if len(alias) <= 3:
            if h == alias:
                return True
        else:
            if h == alias or alias in h:
                return True
    return False


def _resolve_field_name(header: str) -> str | None:
    """Return the canonical FI_Tekninen label for a given column header,
    or ``None`` when the column is unrecognised (and therefore ignored)."""
    for canonical, aliases in _FIELD_ALIASES.items():
        if _match_alias(header, aliases):
            return canonical
    return None


def _expand_slash_header(header: str) -> list[str | None]:
    """Expand a slash-separated combined header that spans multiple
    data columns into one canonical field name per token.

    Lauri's RefDesign Teholuettelo uses headers like
    ``"KYLMÄ-/SÄHKÖ-/VASTUSTEHO [kW]"`` that span three Excel columns
    via merged cells — one cell carries the header text, the two
    following header cells are blank, and the data row carries three
    numeric values (one per power type). Splitting the header on
    ``/`` and matching each token gives us
    ``["Jäähdytysteho", "Sähköteho", "Vastusteho"]``.

    A token alone (e.g. ``"kylmä-"``) often won't match an alias on
    its own — the unit annotation ``[kW]`` carries the implicit
    ``-teho`` suffix. When the unit looks like a power unit and a
    token doesn't already contain ``"teho"`` we retry with
    ``token + "teho"`` appended.

    Returns a list of canonical field names (or ``None`` for tokens
    that didn't resolve). For headers without ``/`` the list has a
    single element, matching the legacy single-column behaviour.
    """
    if "/" not in header:
        return [_resolve_field_name(header)]

    # Split off the unit annotation ("[kW]" / "(kW)" etc.) — we want
    # to keep it for fallback matching but strip it from token text.
    unit = ""
    body = header
    for bracket in ("[", "("):
        idx = header.find(bracket)
        if idx != -1:
            unit = header[idx:]
            body = header[:idx].rstrip()
            break

    raw_tokens = [t.strip().rstrip(" -") for t in body.split("/")]

    is_power_unit = "w" in unit.lower()  # 'kW', 'W'

    result: list[str | None] = []
    for tok in raw_tokens:
        if not tok:
            result.append(None)
            continue
        canonical = _resolve_field_name(tok)
        if canonical is None and is_power_unit and "teho" not in tok.lower():
            canonical = _resolve_field_name(tok + "teho")
        result.append(canonical)
    return result


def _resolve_key_columns(headers: list[str]) -> tuple[int, int] | None:
    """Find the column indices of the koneikko and laitetunnus headers.
    Returns ``None`` when either is missing."""
    koneikko_idx: int | None = None
    laite_idx: int | None = None
    for i, h in enumerate(headers):
        if koneikko_idx is None and _match_alias(h, _KONEIKKO_ALIASES):
            koneikko_idx = i
            continue
        if laite_idx is None and _match_alias(h, _LAITETUNNUS_ALIASES):
            laite_idx = i
    if koneikko_idx is None or laite_idx is None:
        return None
    return koneikko_idx, laite_idx


def _stringify(value: Any) -> str:
    """Render a cell value as the string we'll write into FI_Tekninen.

    Numbers keep their native formatting (no extra decimals when the
    Excel cell was a whole number). ``None`` and empty strings collapse
    to ``""`` so the PSet writer can skip the line.
    """
    if value is None:
        return ""
    if isinstance(value, float):
        # Drop the .0 suffix when the value is integral (5.0 → "5").
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, bool):
        return "kyllä" if value else "ei"
    return str(value).strip()


def _normalise_key(value: Any) -> str:
    """Match-key for koneikko / laitetunnus.

    Both lookup-side (DXF POSITIO TEKSTI/NUMERO) and table-side (Excel
    cell) get the same treatment: stringify, strip, casefold. So
    ``"JK1"`` matches ``"jk1"``, and ``"5"`` matches ``5.0`` in Excel.
    """
    return _stringify(value).strip().casefold()


def _rows_from_csv(path: Path) -> list[tuple[str, list[str], list[list[Any]]]]:
    """Read a CSV file. Auto-detects the dialect (delimiter / quoting).

    Returns a single (sheet_name, headers, body) tuple — CSV has no
    sheets but we keep the list shape so the caller can use one code
    path for both formats.
    """
    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        rows: list[list[Any]] = [list(r) for r in reader if any(c.strip() for c in r)]
    if not rows:
        return [(path.stem, [], [])]
    headers, body = _split_header_and_body(rows)
    return [(path.stem, headers, body)]


def _rows_from_xlsx(path: Path) -> list[tuple[str, list[str], list[list[Any]]]]:
    """Read every sheet from an .xlsx file.

    Lauri's RefDesign teholuettelo splits refrigeration data across
    multiple sheets (Pakasteet for frozen, Kylmät for chilled, Yleiset
    for the overview); reading only ``wb.active`` would miss most of
    the data. Each sheet is parsed independently and rows are merged
    later. Sheets without a recognisable Koneikko + Laitetunnus header
    are returned with empty headers/body and dropped by the caller.
    """
    import openpyxl

    wb = openpyxl.load_workbook(filename=str(path), data_only=True, read_only=True)
    out: list[tuple[str, list[str], list[list[Any]]]] = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        raw_rows: list[list[Any]] = [list(row) for row in sheet.iter_rows(values_only=True)]
        populated = [
            r for r in raw_rows if any(c is not None and str(c).strip() for c in r)
        ]
        if not populated:
            out.append((sheet_name, [], []))
            continue
        headers, body = _split_header_and_body(populated)
        out.append((sheet_name, headers, body))
    wb.close()
    return out


def _split_header_and_body(
    rows: list[list[Any]],
) -> tuple[list[str], list[list[Any]]]:
    """Find the first row that looks like a header (has both a
    koneikko-like and a laitetunnus-like column) and split the table.

    The pre-header rows are PREPENDED to the body so the section-
    header detector (in :func:`_specs_from_table`) can see e.g.
    ``"PAKASTEET JK1"`` that RefDesign templates place one row above
    the column headers. Pre-header title rows are harmless on the
    data side because they have no koneikko/laite columns populated
    and get skipped naturally.

    Tolerates one or more title rows above the table — Lauri's xlsx
    has 6 title rows + a section heading + a unit row before the
    actual data starts.
    """
    header_idx = 0
    for i, row in enumerate(rows):
        candidates = [_stringify(c) for c in row]
        if _resolve_key_columns(candidates) is not None:
            header_idx = i
            break
    headers = [_stringify(c) for c in rows[header_idx]]
    body = rows[:header_idx] + rows[header_idx + 1 :]
    return headers, body


def _specs_from_table(
    headers: list[str], body: list[list[Any]]
) -> dict[tuple[str, str], EnergySpec]:
    """Convert one (headers, body) table into the lookup dict."""
    if not headers:
        return {}
    keys = _resolve_key_columns(headers)
    if keys is None:
        return {}
    koneikko_idx, laite_idx = keys

    field_map: dict[int, str] = {}
    for i, h in enumerate(headers):
        if i in (koneikko_idx, laite_idx):
            continue
        # A slash-separated combined header spans the next N columns.
        # Each token resolves independently; we assign each token to
        # the column at offset ``i + offset``. Skip tokens that don't
        # resolve and skip offsets that would collide with the
        # koneikko/laitetunnus columns or an already-claimed entry.
        expanded = _expand_slash_header(h)
        if len(expanded) > 1:
            for offset, canonical in enumerate(expanded):
                col_idx = i + offset
                if canonical is None:
                    continue
                if col_idx in (koneikko_idx, laite_idx):
                    continue
                if col_idx in field_map:
                    continue
                field_map[col_idx] = canonical
        else:
            canonical = expanded[0]
            if canonical is not None and i not in field_map:
                field_map[i] = canonical

    specs: dict[tuple[str, str], EnergySpec] = {}
    # Forward-filled koneikko: persists across rows where the koneikko
    # cell is blank (Lauri:n RefDesign-pohja jättää REV.-sarakkeen
    # tyhjäksi useimmille riveille; sektio-otsikko ``PAKASTEET JK1``
    # antaa koneikon, ja seuraavien rivien REV. on blank — pitäisi
    # periytyä).
    current_koneikko: str = ""
    current_koneikko_display: str = ""
    for raw_row in body:
        row = list(raw_row) + [None] * (max(len(headers), 0) - len(raw_row))
        # Section-header detection: a row whose only non-empty cells
        # are non-numeric and whose text contains a koneikko-shaped
        # token (JK1, KK2, RK10 …) updates ``current_koneikko``. Such
        # rows have no real laite_value so we move on after capturing.
        section_koneikko = _detect_section_koneikko(row)
        if section_koneikko is not None:
            current_koneikko_display = section_koneikko
            current_koneikko = _normalise_key(section_koneikko)
            continue

        koneikko_value = row[koneikko_idx] if koneikko_idx < len(row) else None
        laite_value = row[laite_idx] if laite_idx < len(row) else None
        laite_key = _normalise_key(laite_value)

        # The koneikko column often carries free text ("Sähköurakoitsija
        # tuo syötöt…") instead of an actual koneikko code in Lauri:n
        # RefDesign-pohjissa. Only update ``current_koneikko`` when the
        # cell text contains a recognised koneikko token (JKx/KKx/RKx/
        # LAx); otherwise forward-fill the section-header / previous
        # value. Empty cells also forward-fill.
        koneikko_text = _stringify(koneikko_value)
        koneikko_match = _KONEIKKO_TOKEN_RE.search(koneikko_text)
        if koneikko_match is not None:
            current_koneikko_display = koneikko_match.group(1).upper()
            current_koneikko = _normalise_key(current_koneikko_display)
        koneikko_key = current_koneikko
        koneikko_display = current_koneikko_display
        if not koneikko_key or not laite_key:
            continue
        fields: dict[str, str] = {}
        for col_idx, canonical in field_map.items():
            if col_idx >= len(row):
                continue
            value_str = _stringify(row[col_idx])
            if value_str:
                fields[canonical] = value_str
        if not fields:
            continue
        specs[(koneikko_key, laite_key)] = EnergySpec(
            koneikko=koneikko_display,
            laitetunnus=_stringify(laite_value),
            fields=fields,
        )
    return specs


# Recognised koneikko-token shapes — RefDesign convention is JK + digits
# (Jäähdytyskoneikko), KK (Kylmäkoneikko), RK (Ryhmäkeskus), LA (Laite).
_KONEIKKO_TOKEN_RE = re.compile(r"\b((?:JK|KK|RK|LA)\d+)\b", re.IGNORECASE)


def _detect_section_koneikko(row: list[Any]) -> str | None:
    """If ``row`` looks like a section header carrying a koneikko code,
    return that code. Otherwise return ``None``.

    Lauri:n RefDesign-pohja merkitsee koneikko-sektioita riveillä jossa
    yksi solu sisältää tekstin kuten ``"PAKASTEET JK1"`` ja muut solut
    ovat tyhjiä. We require: exactly one non-empty cell, the cell is
    text (not numeric/formula), and the text contains a JKx/KKx/RKx/LAx
    token. A pure equipment row (``"KYL-KK-JK1-2"`` in column 1) won't
    match because column-2 (``Koneikkokeskus``) is also populated.
    """
    populated = [c for c in row if c is not None and str(c).strip()]
    if len(populated) != 1:
        return None
    cell = populated[0]
    if isinstance(cell, (int, float)) and not isinstance(cell, bool):
        return None
    text = str(cell)
    match = _KONEIKKO_TOKEN_RE.search(text)
    if match is None:
        return None
    return match.group(1).upper()


def _read_sheets(
    path: Path,
) -> list[tuple[str, list[str], list[list[Any]]]]:
    """Dispatch on file extension and return the per-sheet rows."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _rows_from_xlsx(path)
    if suffix in (".csv", ".tsv", ".txt"):
        return _rows_from_csv(path)
    raise ValueError(
        f"Unsupported energy-spec file extension {suffix!r} "
        f"(supported: .xlsx, .xlsm, .csv, .tsv, .txt)"
    )


def load_energy_specs(path: str | Path) -> dict[tuple[str, str], EnergySpec]:
    """Parse an energy-spec file into a (koneikko, laitetunnus) lookup.

    Supports ``.xlsx`` / ``.xlsm`` (openpyxl) and ``.csv`` / ``.tsv``
    (stdlib). Multi-sheet xlsx is fully supported — every sheet is
    parsed and the rows merged into a single lookup. Sheets without a
    recognisable Koneikko + Laitetunnus header are skipped.
    Returns an empty dict if no sheet has usable rows.
    """
    specs: dict[tuple[str, str], EnergySpec] = {}
    for _name, headers, body in _read_sheets(Path(path)):
        specs.update(_specs_from_table(headers, body))
    return specs


def load_energy_specs_with_headers(
    path: str | Path,
) -> tuple[dict[tuple[str, str], EnergySpec], dict[str, list[str]]]:
    """Same as :func:`load_energy_specs` but also returns the per-sheet
    headers actually parsed.

    Used by the orchestrator's diagnostic logging: when no rows are
    found, we want to show the user which headers WERE detected so
    they can spot the column-name mismatch instead of staring at a
    silent empty result.
    """
    specs: dict[tuple[str, str], EnergySpec] = {}
    headers_per_sheet: dict[str, list[str]] = {}
    for name, headers, body in _read_sheets(Path(path)):
        if headers:
            headers_per_sheet[name] = headers
        specs.update(_specs_from_table(headers, body))
    return specs, headers_per_sheet


def lookup_spec(
    specs: dict[tuple[str, str], EnergySpec],
    *,
    koneikko: str | None,
    laitetunnus: str | None,
) -> EnergySpec | None:
    """Look up a spec by (koneikko, laitetunnus). Tolerates None on
    either side so callers can pass extra_props values directly."""
    if not koneikko or not laitetunnus:
        return None
    key = (_normalise_key(koneikko), _normalise_key(laitetunnus))
    return specs.get(key)
