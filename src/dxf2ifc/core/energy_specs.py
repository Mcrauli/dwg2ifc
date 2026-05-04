"""Read refrigeration-equipment energy specs from CSV / XLSX and produce
a per-(koneikko, laitetunnus) lookup that the orchestrator can merge
into FI_Tekninen.

The Finnish refrigeration workflow keeps energy specs (Jäähdytysteho,
Sähköteho, Kylmäaine, ilmavirta, äänitiedot, käyttölämpötilat) in an
Excel spreadsheet — one row per höyrystin / lauhdutin / kompressori,
keyed by koneikkotunnus + laitetunnus. POSITIO-blokit DXF-pohjassa
kantavat samat tunnukset, joten kun POSITIO-linkitys on kerran tehty,
me voidaan automaattisesti hakea kunkin laitteen oikeat tehotiedot
listalta ja kirjoittaa ne IFC:n FI_Tekninen-PSetiin.

Sarakeotsikoiden mätsäys on tarkoituksella laaja:
``Jäähdytysteho [kW]`` / ``Q_kW`` / ``Cooling capacity`` kaikki
mappautuvat samaksi FI_Tekninen-kentäksi ``Jäähdytysteho``.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable


# Canonical FI_Tekninen field name → list of recognised aliases.
# Matching is case-insensitive against the column header after stripping
# whitespace, brackets, and the unit ("[kW]", "(kW)" etc.).
_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "Jäähdytysteho": (
        "jäähdytysteho",
        "jaahdytysteho",
        "jäähdytys",
        "jaahdytys",
        "jäähdytys teho",
        "kylmäteho",
        "cooling capacity",
        "cooling power",
        "q_kw",
        "q kw",
        "qe",
        "qe_kw",
    ),
    "Lauhdutusteho": (
        "lauhdutusteho",
        "lauhdutus teho",
        "lauhdutus",
        "qc",
        "qc_kw",
        "condensing capacity",
        "condenser power",
    ),
    "Sähköteho": (
        "sähköteho",
        "sahkoteho",
        "sähkö teho",
        "sähkö",
        "sahko",
        "p_kw",
        "p kw",
        "electric power",
        "electrical power",
    ),
    "Kylmäaine": (
        "kylmäaine",
        "kylmaaine",
        "kylmä aine",
        "kylma aine",
        "refrigerant",
        "aine",
    ),
    "Ilmavirta": (
        "ilmavirta",
        "ilma virta",
        "air flow",
        "airflow",
        "air volume",
        "m3/h",
        "m³/h",
    ),
    "Ääniteho": (
        "ääniteho",
        "aaniteho",
        "ääni teho",
        "äänitaso",
        "aanitaso",
        "sound power",
        "noise level",
        "db(a)",
        "lwa",
    ),
    "Käyttölämpötila": (
        "käyttölämpötila",
        "kayttolampotila",
        "käyttö lämpötila",
        "operating temperature",
        "lämpötila",
        "lampotila",
    ),
    "Höyrystymislämpötila": (
        "höyrystymislämpötila",
        "hoyrystymislampotila",
        "evaporation temperature",
        "te",
        "to",
    ),
    "Lauhtumislämpötila": (
        "lauhtumislämpötila",
        "lauhtumislampotila",
        "condensing temperature",
        "tc",
    ),
    "Vastusteho": (
        "vastusteho",
        "sulatusteho",
        "vastus teho",
        "defrost",
        "defrost power",
        "heater",
    ),
    "Jännite": (
        "jännite",
        "jannite",
        "voltage",
        "u_v",
    ),
    "Jäähdyttävä vaikutus": (
        "jäähdyttävä vaikutus",
        "jaahdyttava vaikutus",
        "cooling effect",
    ),
}

# Header aliases for the two key columns. Same case-insensitive matching.
_KONEIKKO_ALIASES: tuple[str, ...] = (
    "koneikko",
    "koneikkotunnus",
    "koneikko-tunnus",
    "koneikko tunnus",
    "koneryhmä",
    "koneryhma",
    "unit",
    "system",
    # RefDesign-konventio Lauri'n teholuetteloissa: "REV." -sarake
    # sisältää koneikkotunnuksen (JK1, JK2…), ei revision-numeroa.
    "rev.",
    "rev",
)
_LAITETUNNUS_ALIASES: tuple[str, ...] = (
    "laitetunnus",
    "laite tunnus",
    "laite-tunnus",
    "laite",
    "tunnus",
    "pos",
    "pos.",
    "positio",
    "positionumero",
    "numero",
    "nr",
    "device",
    "tag",
)


@dataclass(frozen=True)
class EnergySpec:
    """Resolved energy fields for one device. ``fields`` is a
    canonical-name → string-value mapping ready to merge into
    FI_Tekninen."""

    koneikko: str
    laitetunnus: str
    fields: dict[str, str]


def _normalise_header(raw: str) -> str:
    """Lowercase + strip + remove unit brackets / annotations.

    "Jäähdytysteho [kW]" → "jäähdytysteho"
    "Q_kW (cooling)"     → "q_kw"
    "  Laitetunnus  "    → "laitetunnus"
    """
    s = raw or ""
    # Drop everything from the first '(' or '[' onward — that's almost
    # always a unit annotation we already strip implicitly.
    for sep in ("[", "("):
        idx = s.find(sep)
        if idx != -1:
            s = s[:idx]
    return s.strip().lower()


def _match_alias(header: str, aliases: Iterable[str]) -> bool:
    """Return True when the normalised header equals OR contains any alias.

    Short aliases (≤3 characters) require an exact match — substring
    matching for tokens like "te", "tc", "u_v" would otherwise leak
    across distinct fields ("vastusteho" should not match "Te" =
    Evaporation temperature).
    """
    h = _normalise_header(header)
    for alias in aliases:
        if len(alias) <= 3:
            if h == alias:
                return True
        else:
            if h == alias or alias in h:
                return True
    return False


def _resolve_field_name(header: str) -> str | None:
    """Return the canonical FI_Tekninen label for a given column header,
    or ``None`` when the column is unrecognised (and therefore ignored)."""
    for canonical, aliases in _FIELD_ALIASES.items():
        if _match_alias(header, aliases):
            return canonical
    return None


def _resolve_key_columns(headers: list[str]) -> tuple[int, int] | None:
    """Find the column indices of the koneikko and laitetunnus headers.
    Returns ``None`` when either is missing."""
    koneikko_idx: int | None = None
    laite_idx: int | None = None
    for i, h in enumerate(headers):
        if koneikko_idx is None and _match_alias(h, _KONEIKKO_ALIASES):
            koneikko_idx = i
            continue
        if laite_idx is None and _match_alias(h, _LAITETUNNUS_ALIASES):
            laite_idx = i
    if koneikko_idx is None or laite_idx is None:
        return None
    return koneikko_idx, laite_idx


def _stringify(value: Any) -> str:
    """Render a cell value as the string we'll write into FI_Tekninen.

    Numbers keep their native formatting (no extra decimals when the
    Excel cell was a whole number). ``None`` and empty strings collapse
    to ``""`` so the PSet writer can skip the line.
    """
    if value is None:
        return ""
    if isinstance(value, float):
        # Drop the .0 suffix when the value is integral (5.0 → "5").
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, bool):
        return "kyllä" if value else "ei"
    return str(value).strip()


def _normalise_key(value: Any) -> str:
    """Match-key for koneikko / laitetunnus.

    Both lookup-side (DXF POSITIO TEKSTI/NUMERO) and table-side (Excel
    cell) get the same treatment: stringify, strip, casefold. So
    ``"JK1"`` matches ``"jk1"``, and ``"5"`` matches ``5.0`` in Excel.
    """
    return _stringify(value).strip().casefold()


def _rows_from_csv(path: Path) -> list[tuple[str, list[str], list[list[Any]]]]:
    """Read a CSV file. Auto-detects the dialect (delimiter / quoting).

    Returns a single (sheet_name, headers, body) tuple — CSV has no
    sheets but we keep the list shape so the caller can use one code
    path for both formats.
    """
    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        rows: list[list[Any]] = [list(r) for r in reader if any(c.strip() for c in r)]
    if not rows:
        return [(path.stem, [], [])]
    headers, body = _split_header_and_body(rows)
    return [(path.stem, headers, body)]


def _rows_from_xlsx(path: Path) -> list[tuple[str, list[str], list[list[Any]]]]:
    """Read every sheet from an .xlsx file.

    Lauri's RefDesign teholuettelo splits refrigeration data across
    multiple sheets (Pakasteet for frozen, Kylmät for chilled, Yleiset
    for the overview); reading only ``wb.active`` would miss most of
    the data. Each sheet is parsed independently and rows are merged
    later. Sheets without a recognisable Koneikko + Laitetunnus header
    are returned with empty headers/body and dropped by the caller.
    """
    import openpyxl

    wb = openpyxl.load_workbook(filename=str(path), data_only=True, read_only=True)
    out: list[tuple[str, list[str], list[list[Any]]]] = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        raw_rows: list[list[Any]] = [list(row) for row in sheet.iter_rows(values_only=True)]
        populated = [
            r for r in raw_rows if any(c is not None and str(c).strip() for c in r)
        ]
        if not populated:
            out.append((sheet_name, [], []))
            continue
        headers, body = _split_header_and_body(populated)
        out.append((sheet_name, headers, body))
    wb.close()
    return out


def _split_header_and_body(
    rows: list[list[Any]],
) -> tuple[list[str], list[list[Any]]]:
    """Find the first row that looks like a header (has both a
    koneikko-like and a laitetunnus-like column) and split the table.

    Tolerates one or more title rows above the table — Lauri's xlsx
    has 6 title rows + a section heading + a unit row before the
    actual data starts.
    """
    header_idx = 0
    for i, row in enumerate(rows):
        candidates = [_stringify(c) for c in row]
        if _resolve_key_columns(candidates) is not None:
            header_idx = i
            break
    headers = [_stringify(c) for c in rows[header_idx]]
    body = rows[header_idx + 1 :]
    return headers, body


def _specs_from_table(
    headers: list[str], body: list[list[Any]]
) -> dict[tuple[str, str], EnergySpec]:
    """Convert one (headers, body) table into the lookup dict."""
    if not headers:
        return {}
    keys = _resolve_key_columns(headers)
    if keys is None:
        return {}
    koneikko_idx, laite_idx = keys

    field_map: dict[int, str] = {}
    for i, h in enumerate(headers):
        if i in (koneikko_idx, laite_idx):
            continue
        canonical = _resolve_field_name(h)
        if canonical is not None:
            field_map[i] = canonical

    specs: dict[tuple[str, str], EnergySpec] = {}
    for raw_row in body:
        row = list(raw_row) + [None] * (max(len(headers), 0) - len(raw_row))
        koneikko_value = row[koneikko_idx] if koneikko_idx < len(row) else None
        laite_value = row[laite_idx] if laite_idx < len(row) else None
        koneikko_key = _normalise_key(koneikko_value)
        laite_key = _normalise_key(laite_value)
        if not koneikko_key or not laite_key:
            continue
        fields: dict[str, str] = {}
        for col_idx, canonical in field_map.items():
            if col_idx >= len(row):
                continue
            value_str = _stringify(row[col_idx])
            if value_str:
                fields[canonical] = value_str
        if not fields:
            continue
        specs[(koneikko_key, laite_key)] = EnergySpec(
            koneikko=_stringify(koneikko_value),
            laitetunnus=_stringify(laite_value),
            fields=fields,
        )
    return specs


def _read_sheets(
    path: Path,
) -> list[tuple[str, list[str], list[list[Any]]]]:
    """Dispatch on file extension and return the per-sheet rows."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _rows_from_xlsx(path)
    if suffix in (".csv", ".tsv", ".txt"):
        return _rows_from_csv(path)
    raise ValueError(
        f"Unsupported energy-spec file extension {suffix!r} "
        f"(supported: .xlsx, .xlsm, .csv, .tsv, .txt)"
    )


def load_energy_specs(path: str | Path) -> dict[tuple[str, str], EnergySpec]:
    """Parse an energy-spec file into a (koneikko, laitetunnus) lookup.

    Supports ``.xlsx`` / ``.xlsm`` (openpyxl) and ``.csv`` / ``.tsv``
    (stdlib). Multi-sheet xlsx is fully supported — every sheet is
    parsed and the rows merged into a single lookup. Sheets without a
    recognisable Koneikko + Laitetunnus header are skipped.
    Returns an empty dict if no sheet has usable rows.
    """
    specs: dict[tuple[str, str], EnergySpec] = {}
    for _name, headers, body in _read_sheets(Path(path)):
        specs.update(_specs_from_table(headers, body))
    return specs


def load_energy_specs_with_headers(
    path: str | Path,
) -> tuple[dict[tuple[str, str], EnergySpec], dict[str, list[str]]]:
    """Same as :func:`load_energy_specs` but also returns the per-sheet
    headers actually parsed.

    Used by the orchestrator's diagnostic logging: when no rows are
    found, we want to show the user which headers WERE detected so
    they can spot the column-name mismatch instead of staring at a
    silent empty result.
    """
    specs: dict[tuple[str, str], EnergySpec] = {}
    headers_per_sheet: dict[str, list[str]] = {}
    for name, headers, body in _read_sheets(Path(path)):
        if headers:
            headers_per_sheet[name] = headers
        specs.update(_specs_from_table(headers, body))
    return specs, headers_per_sheet


def lookup_spec(
    specs: dict[tuple[str, str], EnergySpec],
    *,
    koneikko: str | None,
    laitetunnus: str | None,
) -> EnergySpec | None:
    """Look up a spec by (koneikko, laitetunnus). Tolerates None on
    either side so callers can pass extra_props values directly."""
    if not koneikko or not laitetunnus:
        return None
    key = (_normalise_key(koneikko), _normalise_key(laitetunnus))
    return specs.get(key)
