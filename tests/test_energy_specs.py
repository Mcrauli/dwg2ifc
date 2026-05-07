"""Tests for the energy-spec loader (Excel + CSV → FI_Tekninen lookup)."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from dxf2ifc.core.energy_specs import (
    EnergySpec,
    load_energy_specs,
    load_energy_specs_with_headers,
    lookup_spec,
)


def _write_csv(path: Path, rows: list[list[str]]) -> None:
    import csv

    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)


def _write_xlsx(path: Path, rows: list[list[object]]) -> None:
    wb = openpyxl.Workbook()
    sheet = wb.active
    for row in rows:
        sheet.append(row)
    wb.save(str(path))


class TestLoadEnergySpecsCSV:
    def test_basic_lookup(self, tmp_path: Path) -> None:
        path = tmp_path / "specs.csv"
        _write_csv(
            path,
            [
                ["Koneikko", "Laitetunnus", "Jäähdytysteho [kW]", "Sähköteho [kW]"],
                ["JK1", "5", "5.2", "0.45"],
                ["JK2", "12", "8.1", "0.7"],
            ],
        )
        specs = load_energy_specs(path)
        assert ("jk1", "5") in specs
        spec = specs[("jk1", "5")]
        assert spec.fields == {"Jäähdytysteho": "5.2", "Sähköteho": "0.45"}
        assert spec.koneikko == "JK1"
        assert spec.laitetunnus == "5"

    def test_handles_semicolon_delimiter(self, tmp_path: Path) -> None:
        path = tmp_path / "specs.csv"
        # European Excel default — sniffer should pick up the semicolon.
        path.write_text(
            "Koneikko;Laitetunnus;Jäähdytysteho [kW]\n"
            "JK1;5;5,2\n",
            encoding="utf-8",
        )
        specs = load_energy_specs(path)
        assert ("jk1", "5") in specs

    def test_empty_rows_skipped(self, tmp_path: Path) -> None:
        path = tmp_path / "specs.csv"
        _write_csv(
            path,
            [
                ["Koneikko", "Laitetunnus", "Jäähdytysteho"],
                [],
                ["JK1", "5", "5.2"],
                ["", "", ""],
                ["JK2", "12", "8.1"],
            ],
        )
        specs = load_energy_specs(path)
        assert len(specs) == 2

    def test_alias_columns_resolved(self, tmp_path: Path) -> None:
        # Common variants the user might type.
        path = tmp_path / "specs.csv"
        _write_csv(
            path,
            [
                ["Koneryhmä", "Pos.", "Cooling capacity [kW]", "Refrigerant"],
                ["JK1", "5", "5.2", "R454C"],
            ],
        )
        specs = load_energy_specs(path)
        spec = specs[("jk1", "5")]
        assert spec.fields == {
            "Jäähdytysteho": "5.2",
            "Kylmäaine": "R454C",
        }


class TestLoadEnergySpecsXLSX:
    def test_basic_xlsx(self, tmp_path: Path) -> None:
        path = tmp_path / "specs.xlsx"
        _write_xlsx(
            path,
            [
                ["Koneikko", "Laitetunnus", "Jäähdytysteho [kW]", "Kylmäaine"],
                ["JK1", 5, 5.2, "R454C"],
                ["JK2", 12, 8.1, "R454C"],
            ],
        )
        specs = load_energy_specs(path)
        spec = specs[("jk1", "5")]
        # Numeric laitetunnus 5 in Excel matches "5" string from POSITIO.
        assert spec.fields == {"Jäähdytysteho": "5.2", "Kylmäaine": "R454C"}

    def test_xlsx_skips_title_rows(self, tmp_path: Path) -> None:
        # User often has a title row (or two) above the table.
        path = tmp_path / "specs.xlsx"
        _write_xlsx(
            path,
            [
                ["Energialista 2026"],
                [None],
                ["Koneikko", "Laitetunnus", "Jäähdytysteho", "Kylmäaine"],
                ["JK1", 5, 5.2, "R454C"],
            ],
        )
        specs = load_energy_specs(path)
        assert ("jk1", "5") in specs

    def test_xlsx_integer_floats_render_clean(self, tmp_path: Path) -> None:
        path = tmp_path / "specs.xlsx"
        _write_xlsx(
            path,
            [
                ["Koneikko", "Laitetunnus", "Sähköteho"],
                ["JK1", 5, 1.0],  # Excel often stores ints as floats
            ],
        )
        spec = load_energy_specs(path)[("jk1", "5")]
        assert spec.fields["Sähköteho"] == "1"


class TestKeyDetection:
    def test_missing_keys_returns_empty(self, tmp_path: Path) -> None:
        path = tmp_path / "no_keys.csv"
        _write_csv(
            path,
            [
                ["Some", "Other", "Columns"],
                ["a", "b", "c"],
            ],
        )
        assert load_energy_specs(path) == {}

    def test_blank_key_rows_skipped(self, tmp_path: Path) -> None:
        path = tmp_path / "blank.csv"
        _write_csv(
            path,
            [
                ["Koneikko", "Laitetunnus", "Jäähdytysteho"],
                ["", "", "5.2"],
                ["JK1", "", "5.2"],
                ["JK1", "5", "5.2"],
            ],
        )
        specs = load_energy_specs(path)
        assert list(specs.keys()) == [("jk1", "5")]

    def test_unsupported_extension_raises(self, tmp_path: Path) -> None:
        path = tmp_path / "x.json"
        path.write_text("{}")
        with pytest.raises(ValueError, match="Unsupported"):
            load_energy_specs(path)


class TestLauriExcelFormat:
    """Tests for Lauri's actual RefDesign teholuettelo layout:
    multiple sheets (Pakasteet/Kylmät), title rows above the table,
    a section heading, a unit row, and the koneikko-tunnus living in
    the column labelled 'REV.' (not a real revision column)."""

    def test_rev_column_recognised_as_koneikko(self) -> None:
        from dxf2ifc.core.energy_specs import _resolve_key_columns

        headers = ["REV.", "POS.", "NIMI", "Kylmäteho"]
        keys = _resolve_key_columns(headers)
        assert keys == (0, 1)

    def test_full_refdesign_layout(self, tmp_path: Path) -> None:
        # Mirrors testi luettelo.xlsx structure: 6 metadata rows,
        # section heading, header row, unit row, then data.
        path = tmp_path / "specs.xlsx"
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Pakasteet"
        sheet.append(["Refdesign Oy", None])
        sheet.append([None])
        sheet.append([None])
        sheet.append([None])
        sheet.append([None])
        sheet.append([None])
        sheet.append(["PAKASTEET JK1"])
        sheet.append(
            ["REV.", "POS.", "NIMI", "Kylmäteho", "Sähköteho", "Vastusteho", "JÄNNITE", "Jäähdyttävä vaikutus"]
        )
        sheet.append([None, None, None, "[kW]", "[kW]", "[kW]"])
        sheet.append(["JK1", 1, "Pakastehuone", 4.5, 0.3, 5.3, 230, 0.27])
        sheet.append(["JK1", 2, "Pakastehuone", 4.5, 0.3, 5.3, 230, 0.27])
        sheet.append([0])  # bogus row — first cell numeric 0, rest empty
        wb.save(str(path))

        specs = load_energy_specs(path)
        assert len(specs) == 2
        spec = specs[("jk1", "1")]
        assert spec.fields == {
            "Jäähdytysteho": "4.5",  # via "Kylmäteho" alias
            "Sähköteho": "0.3",
            "Vastusteho": "5.3",
            "Jännite": "230",
            "Jäähdyttävä vaikutus": "0.27",
        }

    def test_multi_sheet_xlsx(self, tmp_path: Path) -> None:
        path = tmp_path / "specs.xlsx"
        wb = openpyxl.Workbook()
        sheet1 = wb.active
        sheet1.title = "Pakasteet"
        sheet1.append(["REV.", "POS.", "Kylmäteho"])
        sheet1.append(["JK1", 1, 5.2])

        sheet2 = wb.create_sheet("Kylmät")
        sheet2.append(["REV.", "POS.", "Kylmäteho"])
        sheet2.append(["JK2", 10, 8.0])
        wb.save(str(path))

        specs = load_energy_specs(path)
        assert ("jk1", "1") in specs
        assert ("jk2", "10") in specs

    def test_yleiset_overview_sheet_skipped(self, tmp_path: Path) -> None:
        # First sheet has no recognisable table → must be skipped, the
        # data sheet's rows must still come through.
        path = tmp_path / "specs.xlsx"
        wb = openpyxl.Workbook()
        s1 = wb.active
        s1.title = "Yleiset"
        s1.append(["Project info", "value"])
        s1.append(["Site", "Kauhajoki"])

        s2 = wb.create_sheet("Pakasteet")
        s2.append(["REV.", "POS.", "Kylmäteho"])
        s2.append(["JK1", 1, 5.2])
        wb.save(str(path))

        specs, headers = load_energy_specs_with_headers(path)
        assert ("jk1", "1") in specs
        # Yleiset's non-table headers may still appear in the diagnostic
        # output, but the spec lookup itself must be clean.
        assert len(specs) == 1


class TestLoadWithHeaders:
    def test_returns_per_sheet_headers(self, tmp_path: Path) -> None:
        path = tmp_path / "specs.xlsx"
        wb = openpyxl.Workbook()
        s1 = wb.active
        s1.title = "Pakasteet"
        s1.append(["REV.", "POS.", "Kylmäteho", "Sähköteho"])
        s1.append(["JK1", 1, 5.2, 0.3])
        wb.save(str(path))

        specs, headers = load_energy_specs_with_headers(path)
        assert "Pakasteet" in headers
        assert headers["Pakasteet"] == ["REV.", "POS.", "Kylmäteho", "Sähköteho"]

    def test_no_table_returns_empty_with_headers_per_sheet(
        self, tmp_path: Path
    ) -> None:
        # A file the user might paste data into without setting up the
        # right column names. Headers ARE detected (even if useless)
        # so the diagnostic message can quote them back.
        path = tmp_path / "specs.csv"
        path.write_text("Foo,Bar,Baz\n1,2,3\n", encoding="utf-8")
        specs, headers = load_energy_specs_with_headers(path)
        assert specs == {}
        assert "specs" in headers  # CSV uses path stem as sheet name
        assert headers["specs"] == ["Foo", "Bar", "Baz"]


class TestNewFieldAliases:
    def test_vastusteho_alias(self, tmp_path: Path) -> None:
        path = tmp_path / "v.csv"
        path.write_text(
            "Koneikko,Laitetunnus,Vastusteho\nJK1,5,5.3\n", encoding="utf-8"
        )
        spec = load_energy_specs(path)[("jk1", "5")]
        assert spec.fields == {"Vastusteho": "5.3"}

    def test_jannite_alias(self, tmp_path: Path) -> None:
        path = tmp_path / "v.csv"
        path.write_text(
            "Koneikko,Laitetunnus,JÄNNITE\nJK1,5,400\n", encoding="utf-8"
        )
        spec = load_energy_specs(path)[("jk1", "5")]
        assert spec.fields == {"Jännite": "400"}

    def test_jaahdyttava_vaikutus_alias(self, tmp_path: Path) -> None:
        path = tmp_path / "v.csv"
        path.write_text(
            "Koneikko,Laitetunnus,Jäähdyttävä vaikutus\nJK1,5,0.27\n",
            encoding="utf-8",
        )
        spec = load_energy_specs(path)[("jk1", "5")]
        assert spec.fields == {"Jäähdyttävä vaikutus": "0.27"}


class TestLookupSpec:
    def test_case_insensitive_and_whitespace_tolerant(self) -> None:
        specs = {
            ("jk1", "5"): EnergySpec(
                koneikko="JK1", laitetunnus="5", fields={"Jäähdytysteho": "5.2"}
            )
        }
        assert lookup_spec(specs, koneikko="JK1", laitetunnus="5") is not None
        assert lookup_spec(specs, koneikko="jk1", laitetunnus="5") is not None
        assert lookup_spec(specs, koneikko=" Jk1 ", laitetunnus=" 5 ") is not None

    def test_missing_returns_none(self) -> None:
        specs = {
            ("jk1", "5"): EnergySpec(
                koneikko="JK1", laitetunnus="5", fields={"Jäähdytysteho": "5.2"}
            )
        }
        assert lookup_spec(specs, koneikko="JK1", laitetunnus="9") is None
        assert lookup_spec(specs, koneikko=None, laitetunnus="5") is None
        assert lookup_spec(specs, koneikko="JK1", laitetunnus=None) is None


class TestSlashSeparatedHeader:
    def test_three_powers_in_one_header(self, tmp_path: Path) -> None:
        """RefDesign template uses ``KYLMÄ-/SÄHKÖ-/VASTUSTEHO [kW]`` to
        span three columns. Each token must resolve to its own canonical
        FI_Tekninen field (Jäähdytysteho / Sähköteho / Vastusteho)."""
        path = tmp_path / "specs.xlsx"
        _write_xlsx(
            path,
            [
                ["REV.", "POS.", "KYLMÄ-/SÄHKÖ-/VASTUSTEHO [kW]", None, None],
                ["JK1", "5", 4.0, 0.3, 3.7],
            ],
        )
        specs = load_energy_specs(path)
        assert ("jk1", "5") in specs
        spec = specs[("jk1", "5")]
        assert spec.fields["Jäähdytysteho"] == "4"
        assert spec.fields["Sähköteho"] == "0.3"
        assert spec.fields["Vastusteho"] == "3.7"


class TestSectionHeaderAndForwardFill:
    def test_pakasteet_section_propagates_jk1(self, tmp_path: Path) -> None:
        """``PAKASTEET JK1`` section row above the column headers must
        propagate as the koneikko for every body row whose REV. column
        is blank — RefDesign convention."""
        path = tmp_path / "specs.xlsx"
        _write_xlsx(
            path,
            [
                ["", "PAKASTEET JK1", "", ""],
                ["REV.", "POS.", "Jäähdytysteho [kW]", "Sähköteho [kW]"],
                [None, "1", 4.0, 0.3],
                [None, "2", 2.5, 0.2],
            ],
        )
        specs = load_energy_specs(path)
        assert ("jk1", "1") in specs
        assert ("jk1", "2") in specs
        assert specs[("jk1", "1")].koneikko == "JK1"

    def test_section_change_switches_koneikko(self, tmp_path: Path) -> None:
        """Two sections in one sheet (``PAKASTEET JK1`` + ``KYLMÄT JK2``)
        must switch the active koneikko mid-table."""
        path = tmp_path / "specs.xlsx"
        _write_xlsx(
            path,
            [
                ["", "PAKASTEET JK1", "", ""],
                ["REV.", "POS.", "Jäähdytysteho [kW]", "Sähköteho [kW]"],
                [None, "1", 4.0, 0.3],
                ["", "KYLMÄT JK2", "", ""],
                [None, "20", 5.0, 0.2],
            ],
        )
        specs = load_energy_specs(path)
        assert ("jk1", "1") in specs
        assert ("jk2", "20") in specs

    def test_random_text_in_koneikko_column_does_not_clobber(
        self, tmp_path: Path
    ) -> None:
        """A free-text cell in the REV. column ('Sähköurakoitsija…')
        must NOT become the koneikko — only JK/KK/RK/LA + digits are
        accepted as valid koneikko codes."""
        path = tmp_path / "specs.xlsx"
        _write_xlsx(
            path,
            [
                ["", "PAKASTEET JK1", "", ""],
                ["REV.", "POS.", "Jäähdytysteho [kW]", "Sähköteho [kW]"],
                [None, "1", 4.0, 0.3],
                ["Sähköurakoitsija tuo syötöt", None, None, None],
                [None, "2", 2.5, 0.2],
            ],
        )
        specs = load_energy_specs(path)
        assert specs[("jk1", "1")].koneikko == "JK1"
        # Row 2 must stay JK1 — the free-text row didn't switch koneikko.
        assert specs[("jk1", "2")].koneikko == "JK1"
